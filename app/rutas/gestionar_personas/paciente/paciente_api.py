from flask import Blueprint, request, jsonify, current_app as app
from app.dao.gestionar_personas.paciente.PacienteDao import PacienteDao
from flask import send_file
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfgen import canvas
from openpyxl import Workbook
import io
from datetime import datetime


pacienteapi = Blueprint('pacienteapi', __name__)


# ============================================
# GENERAR PDF DE PACIENTE
# ============================================
@pacienteapi.route('/pacientes/<int:pac_id>/pdf', methods=['GET'])
def generarPDF(pac_id):
    """Genera un PDF profesional con la ficha del paciente"""
    pacientedao = PacienteDao()
    
    try:
        paciente = pacientedao.getPacienteById(pac_id)
        
        if not paciente:
            return jsonify({'success': False, 'error': 'Paciente no encontrado'}), 404
        
        # Crear PDF en memoria
        buffer = io.BytesIO()
        
        # Configurar documento con márgenes profesionales
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=40,
            leftMargin=40,
            topMargin=60,
            bottomMargin=50,
            title="Ficha de Paciente",
            author="Sistema Clínico"
        )
        
        # Obtener estilos
        styles = getSampleStyleSheet()
        
        # Configurar estilos personalizados
        titulo_style = ParagraphStyle(
            'TituloCustom',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        subtitulo_style = ParagraphStyle(
            'SubtituloCustom',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2563eb'),
            spaceAfter=12,
            spaceBefore=16,
            fontName='Helvetica-Bold',
            backColor=colors.HexColor('#eff6ff'),
            borderPadding=8,
            borderWidth=1,
            borderColor=colors.HexColor('#3b82f6')
        )
        
        texto_style = ParagraphStyle(
            'TextoCustom',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=6,
            fontName='Helvetica',
            textColor=colors.HexColor('#1f2937')
        )
        
        # Lista de elementos del documento
        elementos = []
        
        # Título principal
        nombre_completo = f"{paciente.get('nombre', '')} {paciente.get('apellido', '')}".strip()
        elementos.append(Paragraph("FICHA DE PACIENTE", titulo_style))
        elementos.append(Spacer(1, 0.2*inch))
        
        # Tabla de información principal con diseño moderno
        datos_principales = [
            ['Historia Clínica', paciente.get('historia_clinica', 'N/A')],
            ['Nombre Completo', nombre_completo or 'N/A'],
            ['Cédula de Identidad', paciente.get('cedula', 'N/A')],
            ['Fecha de Nacimiento', paciente.get('fecha_nacimiento', 'N/A')],
            ['Edad', f"{paciente.get('edad', 'N/A')} años" if paciente.get('edad') else 'N/A'],
            ['Menor de Edad', 'Sí' if paciente.get('es_menor') else 'No'],
        ]
        
        tabla_principal = Table(datos_principales, colWidths=[2.5*inch, 4*inch])
        tabla_principal.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
            ('BACKGROUND', (1, 0), (1, -1), colors.white),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1f2937')),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#374151')),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elementos.append(tabla_principal)
        elementos.append(Spacer(1, 0.3*inch))
        
        # Sección: Información Personal
        elementos.append(Paragraph("INFORMACIÓN PERSONAL", subtitulo_style))
        
        datos_personales = [
            ['Género', paciente.get('genero', 'N/A')],
            ['Estado Civil', paciente.get('estado_civil', 'N/A')],
            ['Teléfono', paciente.get('telefono', 'N/A')],
            ['Correo Electrónico', paciente.get('correo', 'N/A')],
            ['Domicilio', paciente.get('domicilio', 'N/A')],
            ['Ciudad', paciente.get('ciudad', 'N/A')],
            ['Nivel de Instrucción', paciente.get('nivel_instruccion', 'N/A')],
            ['Profesión', paciente.get('profesion', 'N/A')],
        ]
        
        tabla_personal = Table(datos_personales, colWidths=[2.5*inch, 4*inch])
        tabla_personal.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
            ('BACKGROUND', (1, 0), (1, -1), colors.white),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1f2937')),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#374151')),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elementos.append(tabla_personal)
        elementos.append(Spacer(1, 0.3*inch))
        
        # Sección: Datos del Tutor (si es menor)
        if paciente.get('es_menor') and (paciente.get('nom_madre') or paciente.get('nom_padre')):
            elementos.append(Paragraph("INFORMACIÓN DEL TUTOR", subtitulo_style))
            
            datos_tutor = []
            if paciente.get('nom_madre'):
                datos_tutor.append(['Madre', f"{paciente.get('nom_madre', 'N/A')} - Tel: {paciente.get('tel_madre', 'N/A')}"])
            if paciente.get('nom_padre'):
                datos_tutor.append(['Padre', f"{paciente.get('nom_padre', 'N/A')} - Tel: {paciente.get('tel_padre', 'N/A')}"])
            if paciente.get('colegio'):
                datos_tutor.append(['Colegio', f"{paciente.get('colegio', 'N/A')} - Tel: {paciente.get('tel_colegio', 'N/A')}"])
            
            if datos_tutor:
                tabla_tutor = Table(datos_tutor, colWidths=[2.5*inch, 4*inch])
                tabla_tutor.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#fef3c7')),
                    ('BACKGROUND', (1, 0), (1, -1), colors.HexColor('#fffbeb')),
                    ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#92400e')),
                    ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#78350f')),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#fde68a')),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                elementos.append(tabla_tutor)
                elementos.append(Spacer(1, 0.3*inch))
        
        # Sección: Observaciones
        if paciente.get('observaciones'):
            elementos.append(Paragraph("OBSERVACIONES", subtitulo_style))
            elementos.append(Paragraph(paciente.get('observaciones', ''), texto_style))
            elementos.append(Spacer(1, 0.2*inch))
        
        # Pie de página con fecha de generación
        fecha_generacion = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        elementos.append(Spacer(1, 0.3*inch))
        elementos.append(Paragraph(
            f"<i>Documento generado el {fecha_generacion}</i>",
            ParagraphStyle('Pie', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#6b7280'), alignment=TA_CENTER)
        ))
        
        # Construir PDF
        doc.build(elementos)
        
        buffer.seek(0)
        nombre_archivo = f"paciente_{paciente.get('historia_clinica', pac_id)}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return send_file(buffer, as_attachment=True, download_name=nombre_archivo, mimetype='application/pdf')
        
    except Exception as e:
        app.logger.error(f"Error al generar PDF: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'error': 'Error al generar PDF'}), 500


# ============================================
# GENERAR EXCEL DE PACIENTE
# ============================================
@pacienteapi.route('/pacientes/<int:pac_id>/excel', methods=['GET'])
def generarExcel(pac_id):
    """Genera un archivo Excel con los datos del paciente"""
    pacientedao = PacienteDao()
    
    try:
        paciente = pacientedao.getPacienteById(pac_id)
        
        if not paciente:
            return jsonify({'success': False, 'error': 'Paciente no encontrado'}), 404
        
        # Crear Excel en memoria
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos del Paciente"
        
        # Encabezados
        ws['A1'] = "Campo"
        ws['B1'] = "Valor"
        
        # Datos básicos
        datos = [
            ["Historia Clínica", paciente.get('historia_clinica', 'N/A')],
            ["Nombre", paciente.get('nombre', '')],
            ["Apellido", paciente.get('apellido', '')],
            ["Cédula", paciente.get('cedula', 'N/A')],
            ["Fecha Nacimiento", paciente.get('fecha_nacimiento', 'N/A')],
            ["Edad", f"{paciente.get('edad', 'N/A')} años"],
            ["Es menor de edad", 'Sí' if paciente.get('es_menor') else 'No'],
            ["Género", paciente.get('genero', 'N/A')],
            ["Estado Civil", paciente.get('estado_civil', 'N/A')],
            ["Teléfono", paciente.get('telefono', 'N/A')],
            ["Correo", paciente.get('correo', 'N/A')],
            ["Domicilio", paciente.get('domicilio', 'N/A')],
            ["Ciudad", paciente.get('ciudad', 'N/A')],
            ["Ciudad Nacimiento", paciente.get('ciudad_nacimiento', 'N/A')],
            ["Nivel Instrucción", paciente.get('nivel_instruccion', 'N/A')],
            ["Profesión", paciente.get('profesion', 'N/A')],
        ]
        
        # Agregar datos del tutor si es menor
        if paciente.get('es_menor'):
            datos.extend([
                ["--- Datos del Tutor ---", ""],
                ["Nombre Madre", paciente.get('nom_madre', 'N/A')],
                ["Teléfono Madre", paciente.get('tel_madre', 'N/A')],
                ["Nombre Padre", paciente.get('nom_padre', 'N/A')],
                ["Teléfono Padre", paciente.get('tel_padre', 'N/A')],
                ["Educación", paciente.get('educacion', 'N/A')],
                ["Colegio", paciente.get('colegio', 'N/A')],
                ["Teléfono Colegio", paciente.get('tel_colegio', 'N/A')],
            ])
        
        if paciente.get('observaciones'):
            datos.append(["Observaciones", paciente.get('observaciones', '')])
        
        for idx, (campo, valor) in enumerate(datos, start=2):
            ws[f'A{idx}'] = campo
            ws[f'B{idx}'] = valor
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(buffer, as_attachment=True, download_name=f"paciente_{pac_id}.xlsx", 
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except Exception as e:
        app.logger.error(f"Error al generar Excel: {str(e)}")
        return jsonify({'success': False, 'error': 'Error al generar Excel'}), 500


# ============================================
# OBTENER TODOS LOS PACIENTES
# ============================================
@pacienteapi.route('/pacientes/debug', methods=['GET'])
def debugPacientes():
    """Ruta de debug temporal para verificar sesión y filtros"""
    from flask import session
    from app.utils.especialista_helper import obtener_id_especialista_usuario, puede_ver_todos_pacientes
    
    return jsonify({
        'success': True,
        'debug': {
            'id_usuario': session.get('id_usuario'),
            'id_grupo': session.get('id_grupo'),
            'grupo': session.get('grupo'),
            'puede_ver_todos': puede_ver_todos_pacientes(),
            'id_especialista': obtener_id_especialista_usuario(),
            'sesion_completa': dict(session)
        }
    }), 200


@pacienteapi.route('/pacientes', methods=['GET'])
def getPacientes():
    """
    Obtiene la lista de pacientes con paginación opcional.
    Parámetros de query string:
        - pagina: Número de página (opcional, si no se pasa retorna todos sin paginación)
        - por_pagina: Cantidad de registros por página (opcional, si no se pasa retorna todos)
    
    Si no se pasan parámetros de paginación, retorna array directo para compatibilidad con frontend.
    Si se pasan parámetros, retorna objeto con estructura de paginación.
    """
    pacientedao = PacienteDao()
    
    try:
        # Verificar si se están solicitando parámetros de paginación
        pagina = request.args.get('pagina', None, type=int)
        por_pagina = request.args.get('por_pagina', None, type=int)
        
        # Si no se pasan parámetros de paginación, retornar todos los registros (compatibilidad)
        if pagina is None and por_pagina is None:
            # Usar un valor grande para obtener todos los registros
            resultado = pacientedao.getPacientes(pagina=1, por_pagina=10000)
            # Retornar solo el array de datos para compatibilidad con frontend existente
            return jsonify({
                'success': True,
                'data': resultado['datos'],  # Retornar array directo
                'error': None
            }), 200
        else:
            # Si se pasan parámetros, usar paginación y retornar estructura completa
            pagina = pagina if pagina is not None else 1
            por_pagina = por_pagina if por_pagina is not None else 50
            
            resultado = pacientedao.getPacientes(pagina=pagina, por_pagina=por_pagina)
            
            return jsonify({
                'success': True,
                'data': resultado,  # Retornar objeto con paginación
                'error': None
            }), 200
    
    except Exception as e:
        app.logger.error(f"Error al obtener todos los pacientes: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno. Consulte con el administrador.'
        }), 500


# ============================================
# OBTENER SOLO PACIENTES MENORES
# ============================================
@pacienteapi.route('/pacientes/menores', methods=['GET'])
def getPacientesMenores():
    """Obtiene solo los pacientes menores de edad (calculado automáticamente)"""
    pacientedao = PacienteDao()
    
    try:
        menores = pacientedao.getPacientesMenores()
        
        return jsonify({
            'success': True,
            'data': menores,
            'error': None
        }), 200
    
    except Exception as e:
        app.logger.error(f"Error al obtener pacientes menores: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno. Consulte con el administrador.'
        }), 500


# ============================================
# OBTENER PACIENTE POR ID
# ============================================
@pacienteapi.route('/pacientes/<int:pac_id>', methods=['GET'])
def getPaciente(pac_id):
    """Obtiene un paciente específico por su ID"""
    pacientedao = PacienteDao()
    
    try:
        paciente = pacientedao.getPacienteById(pac_id)
        
        if paciente:
            return jsonify({
                'success': True,
                'data': paciente,
                'error': None
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': 'No se encontró el paciente con el ID proporcionado.'
            }), 404
    
    except Exception as e:
        app.logger.error(f"Error al obtener el paciente: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno. Consulte con el administrador.'
        }), 500


# ============================================
# OBTENER PACIENTE PARA EDITAR
# ============================================
@pacienteapi.route('/pacientes/<int:pac_id>/editar', methods=['GET'])
def getPacienteParaEditar(pac_id):
    """Obtiene paciente con IDs originales para formulario de edición"""
    pacientedao = PacienteDao()

    try:
        paciente = pacientedao.getPacienteParaEditar(pac_id)

        if paciente:
            return jsonify({
                'success': True,
                'data': paciente,
                'error': None
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': 'No se encontró el paciente.'
            }), 404

    except Exception as e:
        app.logger.error(f"Error al obtener paciente para editar: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno.'
        }), 500


# ============================================
# CREAR NUEVO PACIENTE
# ============================================
@pacienteapi.route('/pacientes', methods=['POST'])
def addPaciente():
    """
    Crea un nuevo paciente con todos sus datos.
    """
    data = request.get_json()
    pacientedao = PacienteDao()

    # ✅ CORRECCIÓN: Solo estos 4 campos son REALMENTE obligatorios
    campos_requeridos = ['nombre', 'apellido', 'cedula', 'fecha_nacimiento']

    # Validar campos obligatorios
    for campo in campos_requeridos:
        if campo not in data or not data[campo]:
            return jsonify({
                'success': False,
                'error': f'El campo {campo} es obligatorio y no puede estar vacío.'
            }), 400

    try:
        paciente_id = pacientedao.guardarPaciente(
            # Obligatorios
            nombre=data['nombre'],
            apellido=data['apellido'],
            cedula=data['cedula'],
            fecha_nacimiento=data['fecha_nacimiento'],
            
            # ✅ Ahora telefono también es opcional (como género/ciudad)
            telefono=data.get('telefono'),
            
            # Opcionales (ya estaban bien)
            id_genero=data.get('id_genero'),
            id_estado_civil=data.get('id_estado_civil'),
            correo=data.get('correo'),
            domicilio=data.get('domicilio'),
            id_ciudad=data.get('id_ciudad'),
            id_ciudad_nacimiento=data.get('id_ciudad_nacimiento'),
            id_nivel_instruccion=data.get('id_nivel_instruccion'),
            id_profesion=data.get('id_profesion'),
            
            # Datos de paciente
            historia_clinica=data.get('historia_clinica'),
            observaciones=data.get('observaciones'),
            
            # Datos del menor
            nom_madre=data.get('nom_madre'),
            tel_madre=data.get('tel_madre'),
            nom_padre=data.get('nom_padre'),
            tel_padre=data.get('tel_padre'),
            educacion=data.get('educacion'),
            colegio=data.get('colegio'),
            tel_colegio=data.get('tel_colegio')
        )

        if paciente_id is not None:
            return jsonify({
                'success': True,
                'data': {
                    'id_paciente': paciente_id,
                    'mensaje': 'Paciente creado exitosamente'
                },
                'error': None
            }), 201
        else:
            return jsonify({
                'success': False,
                'error': 'No se pudo crear el paciente. Verifique que: 1) La fecha de nacimiento sea válida, 2) Si es menor de edad, proporcione al menos el nombre de la madre o padre, 3) La historia clínica no esté duplicada.'
            }), 400

    except Exception as e:
        app.logger.error(f"Error inesperado al agregar paciente: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno del servidor. Consulte con el administrador.'
        }), 500


# ============================================
# ACTUALIZAR PACIENTE EXISTENTE
# ============================================
@pacienteapi.route('/pacientes/<int:pac_id>', methods=['PUT'])
def updatePaciente(pac_id):
    data = request.get_json()
    pacientedao = PacienteDao()

    # Verificar que existe
    paciente_existente = pacientedao.getPacienteById(pac_id)
    if not paciente_existente:
        return jsonify({
            'success': False,
            'error': 'No se encontró el paciente con el ID proporcionado.'
        }), 404

    # ✅ CORRECCIÓN: Si historia_clinica no viene o está vacía, usar la existente
    historia_clinica = data.get('historia_clinica')
    if not historia_clinica or historia_clinica.strip() == '':
        historia_clinica = paciente_existente.get('historia_clinica')
        if not historia_clinica:
            app.logger.warning(f"Paciente {pac_id} no tiene historia clínica, generando una nueva")
            historia_clinica = pacientedao.generar_historia_clinica_unica(
                data.get('nombre', ''),
                data.get('apellido', '')
            )

    # ✅ CORRECCIÓN: Solo estos campos son obligatorios (historia_clinica ya está manejada arriba)
    campos_requeridos = ['nombre', 'apellido', 'cedula', 'fecha_nacimiento']

    for campo in campos_requeridos:
        if campo not in data or not data[campo]:
            return jsonify({
                'success': False,
                'error': f'El campo {campo} es obligatorio y no puede estar vacío.'
            }), 400

    try:
        resultado = pacientedao.updatePaciente(
            pac_id=pac_id,
            
            # Obligatorios
            nombre=data['nombre'],
            apellido=data['apellido'],
            cedula=data['cedula'],
            fecha_nacimiento=data['fecha_nacimiento'],
            historia_clinica=historia_clinica,
            
            # ✅ Telefono ahora opcional
            telefono=data.get('telefono'),
            
            # Opcionales
            id_genero=data.get('id_genero'),
            id_estado_civil=data.get('id_estado_civil'),
            correo=data.get('correo'),
            domicilio=data.get('domicilio'),
            id_ciudad=data.get('id_ciudad'),
            id_ciudad_nacimiento=data.get('id_ciudad_nacimiento'),
            id_nivel_instruccion=data.get('id_nivel_instruccion'),
            id_profesion=data.get('id_profesion'),
            
            # Datos de paciente
            observaciones=data.get('observaciones'),
            
            # Datos del menor
            nom_madre=data.get('nom_madre'),
            tel_madre=data.get('tel_madre'),
            nom_padre=data.get('nom_padre'),
            tel_padre=data.get('tel_padre'),
            educacion=data.get('educacion'),
            colegio=data.get('colegio'),
            tel_colegio=data.get('tel_colegio')
        )

        if resultado:
            return jsonify({
                'success': True,
                'data': {
                    'id_paciente': pac_id,
                    'mensaje': 'Paciente actualizado exitosamente'
                },
                'error': None
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': 'No se pudo actualizar el paciente.'
            }), 400

    except Exception as e:
        app.logger.error(f"Error al actualizar paciente: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno.'
        }), 500

# ============================================
# ELIMINAR PACIENTE
# ============================================
@pacienteapi.route('/pacientes/<int:pac_id>', methods=['DELETE'])
def deletePaciente(pac_id):
    """
    Elimina un paciente y todos sus datos asociados.
    Eliminación en cascada: pacientes_menores -> pacientes -> personas
    """
    pacientedao = PacienteDao()

    try:
        if pacientedao.deletePaciente(pac_id):
            return jsonify({
                'success': True,
                'mensaje': f'Paciente con ID {pac_id} eliminado correctamente.',
                'error': None
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': 'No se encontró el paciente con el ID proporcionado o no se pudo eliminar.'
            }), 404

    except Exception as e:
        app.logger.error(f"Error al eliminar paciente: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno. Consulte con el administrador.'
        }), 500@pacienteapi.route('/pacientes', methods=['POST'])
def addPaciente():
    """
    Crea un nuevo paciente con todos sus datos.
    """
    data = request.get_json()
    pacientedao = PacienteDao()

    # ✅ CORRECCIÓN: Solo estos 4 campos son REALMENTE obligatorios
    campos_requeridos = ['nombre', 'apellido', 'cedula', 'fecha_nacimiento']

    # Validar campos obligatorios
    for campo in campos_requeridos:
        if campo not in data or not data[campo]:
            return jsonify({
                'success': False,
                'error': f'El campo {campo} es obligatorio y no puede estar vacío.'
            }), 400

    try:
        paciente_id = pacientedao.guardarPaciente(
            # Obligatorios
            nombre=data['nombre'],
            apellido=data['apellido'],
            cedula=data['cedula'],
            fecha_nacimiento=data['fecha_nacimiento'],
            
            # ✅ Ahora telefono también es opcional (como género/ciudad)
            telefono=data.get('telefono'),
            
            # Opcionales (ya estaban bien)
            id_genero=data.get('id_genero'),
            id_estado_civil=data.get('id_estado_civil'),
            correo=data.get('correo'),
            domicilio=data.get('domicilio'),
            id_ciudad=data.get('id_ciudad'),
            id_ciudad_nacimiento=data.get('id_ciudad_nacimiento'),
            id_nivel_instruccion=data.get('id_nivel_instruccion'),
            id_profesion=data.get('id_profesion'),
            
            # Datos de paciente
            historia_clinica=data.get('historia_clinica'),
            observaciones=data.get('observaciones'),
            
            # Datos del menor
            nom_madre=data.get('nom_madre'),
            tel_madre=data.get('tel_madre'),
            nom_padre=data.get('nom_padre'),
            tel_padre=data.get('tel_padre'),
            educacion=data.get('educacion'),
            colegio=data.get('colegio'),
            tel_colegio=data.get('tel_colegio')
        )

        if paciente_id is not None:
            return jsonify({
                'success': True,
                'data': {
                    'id_paciente': paciente_id,
                    'mensaje': 'Paciente creado exitosamente'
                },
                'error': None
            }), 201
        else:
            return jsonify({
                'success': False,
                'error': 'No se pudo crear el paciente. Verifique que: 1) La fecha de nacimiento sea válida, 2) Si es menor de edad, proporcione al menos el nombre de la madre o padre, 3) La historia clínica no esté duplicada.'
            }), 400

    except Exception as e:
        app.logger.error(f"Error inesperado al agregar paciente: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno del servidor. Consulte con el administrador.'
        }), 500