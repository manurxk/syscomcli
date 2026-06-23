from flask import Blueprint, request, jsonify, current_app as app, session, send_file
from app.dao.mantenimiento.personas.paciente.PacienteDao import PacienteDao
from app.auth.utils.decorators import role_required
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER
from openpyxl import Workbook
import io
from datetime import datetime


pacienteapi = Blueprint('pacienteapi', __name__)


# ============================================
# GENERAR PDF DE PACIENTE
# ============================================
@pacienteapi.route('/pacientes/<int:pac_id>/pdf', methods=['GET'])
@role_required("ADMINISTRADOR", "SUPERADMIN", "RECEPCIONISTA")
def generarPDF(pac_id):
    """Genera un PDF profesional con la ficha del paciente"""
    pacientedao = PacienteDao()

    try:
        paciente = pacientedao.getPacienteById(pac_id)

        if not paciente:
            return jsonify({'success': False, 'error': 'Paciente no encontrado'}), 404

        buffer = io.BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=40,
            leftMargin=40,
            topMargin=60,
            bottomMargin=50,
            title="Ficha de Paciente",
            author="Sistema Clínico"
        )

        styles = getSampleStyleSheet()

        titulo_style = ParagraphStyle(
            'TituloCustom',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )

        subtitulo_style = ParagraphStyle(
            'SubtituloCustom',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2563eb'),
            spaceAfter=12,
            spaceBefore=16,
            fontName='Helvetica-Bold',
            backColor=colors.HexColor('#eff6ff'),
            borderPadding=8,
            borderWidth=1,
            borderColor=colors.HexColor('#3b82f6')
        )

        texto_style = ParagraphStyle(
            'TextoCustom',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=6,
            fontName='Helvetica',
            textColor=colors.HexColor('#1f2937')
        )

        elementos = []

        nombre_completo = f"{paciente.get('nombre', '')} {paciente.get('apellido', '')}".strip()
        elementos.append(Paragraph("FICHA DE PACIENTE", titulo_style))
        elementos.append(Spacer(1, 0.2 * inch))

        datos_principales = [
            ['Historia Clínica', paciente.get('historia_clinica') or 'N/A'],
            ['Nombre Completo', nombre_completo or 'N/A'],
            ['Cédula de Identidad', paciente.get('cedula') or 'N/A'],
            ['Fecha de Nacimiento', paciente.get('fecha_nacimiento') or 'N/A'],
            ['Edad', f"{paciente.get('edad')} años" if paciente.get('edad') is not None else 'N/A'],
            ['Menor de Edad', 'Sí' if paciente.get('es_menor') else 'No'],
        ]

        tabla_principal = Table(datos_principales, colWidths=[2.5 * inch, 4 * inch])
        tabla_principal.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
            ('BACKGROUND', (1, 0), (1, -1), colors.white),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1f2937')),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#374151')),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elementos.append(tabla_principal)
        elementos.append(Spacer(1, 0.3 * inch))

        elementos.append(Paragraph("INFORMACIÓN PERSONAL", subtitulo_style))

        datos_personales = [
            ['Género', paciente.get('genero') or 'N/A'],
            ['Estado Civil', paciente.get('estado_civil') or 'N/A'],
            ['Teléfono', paciente.get('telefono') or 'N/A'],
            ['Correo Electrónico', paciente.get('correo') or 'N/A'],
            ['Domicilio', paciente.get('domicilio') or 'N/A'],
            ['Ciudad', paciente.get('ciudad') or 'N/A'],
            ['Nivel de Instrucción', paciente.get('nivel_instruccion') or 'N/A'],
            ['Profesión', paciente.get('profesion') or 'N/A'],
        ]

        tabla_personal = Table(datos_personales, colWidths=[2.5 * inch, 4 * inch])
        tabla_personal.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
            ('BACKGROUND', (1, 0), (1, -1), colors.white),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1f2937')),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#374151')),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elementos.append(tabla_personal)
        elementos.append(Spacer(1, 0.3 * inch))

        if paciente.get('es_menor') and (paciente.get('nom_madre') or paciente.get('nom_padre')):
            elementos.append(Paragraph("INFORMACIÓN DEL TUTOR", subtitulo_style))

            datos_tutor = []
            if paciente.get('nom_madre'):
                datos_tutor.append(['Madre', f"{paciente.get('nom_madre')} - Tel: {paciente.get('tel_madre') or 'N/A'}"])
            if paciente.get('nom_padre'):
                datos_tutor.append(['Padre', f"{paciente.get('nom_padre')} - Tel: {paciente.get('tel_padre') or 'N/A'}"])
            if paciente.get('colegio'):
                datos_tutor.append(['Colegio', f"{paciente.get('colegio')} - Tel: {paciente.get('tel_colegio') or 'N/A'}"])

            if datos_tutor:
                tabla_tutor = Table(datos_tutor, colWidths=[2.5 * inch, 4 * inch])
                tabla_tutor.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#fef3c7')),
                    ('BACKGROUND', (1, 0), (1, -1), colors.HexColor('#fffbeb')),
                    ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#92400e')),
                    ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#78350f')),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#fde68a')),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                elementos.append(tabla_tutor)
                elementos.append(Spacer(1, 0.3 * inch))

        if paciente.get('observaciones'):
            elementos.append(Paragraph("OBSERVACIONES", subtitulo_style))
            elementos.append(Paragraph(paciente.get('observaciones'), texto_style))
            elementos.append(Spacer(1, 0.2 * inch))

        fecha_generacion = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        elementos.append(Spacer(1, 0.3 * inch))
        elementos.append(Paragraph(
            f"<i>Documento generado el {fecha_generacion}</i>",
            ParagraphStyle('Pie', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#6b7280'), alignment=TA_CENTER)
        ))

        doc.build(elementos)

        buffer.seek(0)
        nombre_archivo = f"paciente_{paciente.get('historia_clinica') or pac_id}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return send_file(buffer, as_attachment=True, download_name=nombre_archivo, mimetype='application/pdf')

    except Exception as e:
        app.logger.error(f"Error al generar PDF: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'error': 'Error al generar PDF'}), 500


# ============================================
# GENERAR EXCEL DE PACIENTE
# ============================================
@pacienteapi.route('/pacientes/<int:pac_id>/excel', methods=['GET'])
@role_required("ADMINISTRADOR", "SUPERADMIN", "RECEPCIONISTA")
def generarExcel(pac_id):
    """Genera un archivo Excel con los datos del paciente"""
    pacientedao = PacienteDao()

    try:
        paciente = pacientedao.getPacienteById(pac_id)

        if not paciente:
            return jsonify({'success': False, 'error': 'Paciente no encontrado'}), 404

        wb = Workbook()
        ws = wb.active
        ws.title = "Datos del Paciente"

        ws['A1'] = "Campo"
        ws['B1'] = "Valor"

        datos = [
            ["Historia Clínica", paciente.get('historia_clinica') or 'N/A'],
            ["Nombre", paciente.get('nombre') or ''],
            ["Apellido", paciente.get('apellido') or ''],
            ["Cédula", paciente.get('cedula') or 'N/A'],
            ["Fecha Nacimiento", paciente.get('fecha_nacimiento') or 'N/A'],
            ["Edad", f"{paciente.get('edad')} años" if paciente.get('edad') is not None else 'N/A'],
            ["Es menor de edad", 'Sí' if paciente.get('es_menor') else 'No'],
            ["Género", paciente.get('genero') or 'N/A'],
            ["Estado Civil", paciente.get('estado_civil') or 'N/A'],
            ["Teléfono", paciente.get('telefono') or 'N/A'],
            ["Correo", paciente.get('correo') or 'N/A'],
            ["Domicilio", paciente.get('domicilio') or 'N/A'],
            ["Ciudad", paciente.get('ciudad') or 'N/A'],
            ["Ciudad Nacimiento", paciente.get('ciudad_nacimiento') or 'N/A'],
            ["Nivel Instrucción", paciente.get('nivel_instruccion') or 'N/A'],
            ["Profesión", paciente.get('profesion') or 'N/A'],
        ]

        if paciente.get('es_menor'):
            datos.extend([
                ["--- Datos del Tutor ---", ""],
                ["Nombre Madre", paciente.get('nom_madre') or 'N/A'],
                ["Teléfono Madre", paciente.get('tel_madre') or 'N/A'],
                ["Nombre Padre", paciente.get('nom_padre') or 'N/A'],
                ["Teléfono Padre", paciente.get('tel_padre') or 'N/A'],
                ["Educación", paciente.get('educacion') or 'N/A'],
                ["Colegio", paciente.get('colegio') or 'N/A'],
                ["Teléfono Colegio", paciente.get('tel_colegio') or 'N/A'],
            ])

        if paciente.get('observaciones'):
            datos.append(["Observaciones", paciente.get('observaciones')])

        for idx, (campo, valor) in enumerate(datos, start=2):
            ws[f'A{idx}'] = campo
            ws[f'B{idx}'] = valor

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(buffer, as_attachment=True, download_name=f"paciente_{pac_id}.xlsx",
                          mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        app.logger.error(f"Error al generar Excel: {str(e)}")
        return jsonify({'success': False, 'error': 'Error al generar Excel'}), 500


# ============================================
# OBTENER TODOS LOS PACIENTES
# ============================================
@pacienteapi.route('/pacientes', methods=['GET'])
@role_required("ADMINISTRADOR", "SUPERADMIN", "RECEPCIONISTA")
def getPacientes():
    """
    Obtiene la lista de pacientes con paginación opcional.
    Parámetros de query string:
        - pagina: Número de página (opcional, si no se pasa retorna todos sin paginación)
        - por_pagina: Cantidad de registros por página (opcional, si no se pasa retorna todos)

    Si no se pasan parámetros de paginación, retorna array directo para compatibilidad con frontend.
    Si se pasan parámetros, retorna objeto con estructura de paginación.
    """
    pacientedao = PacienteDao()

    try:
        pagina = request.args.get('pagina', None, type=int)
        por_pagina = request.args.get('por_pagina', None, type=int)

        if pagina is None and por_pagina is None:
            resultado = pacientedao.getPacientes(pagina=1, por_pagina=10000)
            return jsonify({
                'success': True,
                'data': resultado['datos'],
                'error': None
            }), 200
        else:
            pagina = pagina if pagina is not None else 1
            por_pagina = por_pagina if por_pagina is not None else 50

            resultado = pacientedao.getPacientes(pagina=pagina, por_pagina=por_pagina)

            return jsonify({
                'success': True,
                'data': resultado,
                'error': None
            }), 200

    except Exception as e:
        app.logger.error(f"Error al obtener todos los pacientes: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno. Consulte con el administrador.'
        }), 500


# ============================================
# OBTENER SOLO PACIENTES MENORES
# ============================================
@pacienteapi.route('/pacientes/menores', methods=['GET'])
@role_required("ADMINISTRADOR", "SUPERADMIN", "RECEPCIONISTA")
def getPacientesMenores():
    """Obtiene solo los pacientes menores de edad (calculado automáticamente)"""
    pacientedao = PacienteDao()

    try:
        menores = pacientedao.getPacientesMenores()

        return jsonify({
            'success': True,
            'data': menores,
            'error': None
        }), 200

    except Exception as e:
        app.logger.error(f"Error al obtener pacientes menores: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno. Consulte con el administrador.'
        }), 500


# ============================================
# OBTENER PACIENTE POR ID
# ============================================
@pacienteapi.route('/pacientes/<int:pac_id>', methods=['GET'])
@role_required("ADMINISTRADOR", "SUPERADMIN", "RECEPCIONISTA")
def getPaciente(pac_id):
    """Obtiene un paciente específico por su ID"""
    pacientedao = PacienteDao()

    try:
        paciente = pacientedao.getPacienteById(pac_id)

        if paciente:
            return jsonify({
                'success': True,
                'data': paciente,
                'error': None
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': 'No se encontró el paciente con el ID proporcionado.'
            }), 404

    except Exception as e:
        app.logger.error(f"Error al obtener el paciente: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno. Consulte con el administrador.'
        }), 500


# ============================================
# OBTENER PACIENTE PARA EDITAR
# ============================================
@pacienteapi.route('/pacientes/<int:pac_id>/editar', methods=['GET'])
@role_required("ADMINISTRADOR", "SUPERADMIN", "RECEPCIONISTA")
def getPacienteParaEditar(pac_id):
    """Obtiene paciente con IDs originales para formulario de edición"""
    pacientedao = PacienteDao()

    try:
        paciente = pacientedao.getPacienteParaEditar(pac_id)

        if paciente:
            return jsonify({
                'success': True,
                'data': paciente,
                'error': None
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': 'No se encontró el paciente.'
            }), 404

    except Exception as e:
        app.logger.error(f"Error al obtener paciente para editar: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno.'
        }), 500


# ============================================
# CREAR NUEVO PACIENTE
# ============================================
@pacienteapi.route('/pacientes', methods=['POST'])
@role_required("ADMINISTRADOR", "SUPERADMIN", "RECEPCIONISTA")
def addPaciente():
    """Crea un nuevo paciente con todos sus datos."""
    data = request.get_json()
    pacientedao = PacienteDao()

    campos_requeridos = ['nombre', 'apellido', 'cedula', 'fecha_nacimiento']

    for campo in campos_requeridos:
        if campo not in data or not data[campo]:
            return jsonify({
                'success': False,
                'error': f'El campo {campo} es obligatorio y no puede estar vacío.'
            }), 400

    try:
        paciente_id = pacientedao.guardarPaciente(
            nombre=data['nombre'],
            apellido=data['apellido'],
            cedula=data['cedula'],
            fecha_nacimiento=data['fecha_nacimiento'],
            telefono=data.get('telefono'),
            id_genero=data.get('id_genero'),
            id_estado_civil=data.get('id_estado_civil'),
            correo=data.get('correo'),
            domicilio=data.get('domicilio'),
            id_ciudad=data.get('id_ciudad'),
            id_ciudad_nacimiento=data.get('id_ciudad_nacimiento'),
            id_nivel_instruccion=data.get('id_nivel_instruccion'),
            id_profesion=data.get('id_profesion'),
            historia_clinica=data.get('historia_clinica'),
            observaciones=data.get('observaciones'),
            nom_madre=data.get('nom_madre'),
            tel_madre=data.get('tel_madre'),
            nom_padre=data.get('nom_padre'),
            tel_padre=data.get('tel_padre'),
            educacion=data.get('educacion'),
            colegio=data.get('colegio'),
            tel_colegio=data.get('tel_colegio'),
            usuario_creacion=session.get('id_usuario')
        )

        if paciente_id is not None:
            return jsonify({
                'success': True,
                'data': {
                    'id_paciente': paciente_id,
                    'mensaje': 'Paciente creado exitosamente'
                },
                'error': None
            }), 201
        else:
            return jsonify({
                'success': False,
                'error': 'No se pudo crear el paciente.'
            }), 400

    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        app.logger.error(f"Error inesperado al agregar paciente: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno del servidor. Consulte con el administrador.'
        }), 500


# ============================================
# ACTUALIZAR PACIENTE EXISTENTE
# ============================================
@pacienteapi.route('/pacientes/<int:pac_id>', methods=['PUT'])
@role_required("ADMINISTRADOR", "SUPERADMIN", "RECEPCIONISTA")
def updatePaciente(pac_id):
    data = request.get_json()
    pacientedao = PacienteDao()

    paciente_existente = pacientedao.getPacienteById(pac_id)
    if not paciente_existente:
        return jsonify({
            'success': False,
            'error': 'No se encontró el paciente con el ID proporcionado.'
        }), 404

    historia_clinica = data.get('historia_clinica')
    if not historia_clinica or historia_clinica.strip() == '':
        historia_clinica = paciente_existente.get('historia_clinica')
        if not historia_clinica:
            app.logger.warning(f"Paciente {pac_id} no tiene historia clínica, generando una nueva")
            historia_clinica = pacientedao.generar_historia_clinica_unica(
                data.get('nombre', ''),
                data.get('apellido', '')
            )

    campos_requeridos = ['nombre', 'apellido', 'cedula', 'fecha_nacimiento']

    for campo in campos_requeridos:
        if campo not in data or not data[campo]:
            return jsonify({
                'success': False,
                'error': f'El campo {campo} es obligatorio y no puede estar vacío.'
            }), 400

    try:
        resultado = pacientedao.updatePaciente(
            pac_id=pac_id,
            nombre=data['nombre'],
            apellido=data['apellido'],
            cedula=data['cedula'],
            fecha_nacimiento=data['fecha_nacimiento'],
            historia_clinica=historia_clinica,
            telefono=data.get('telefono'),
            id_genero=data.get('id_genero'),
            id_estado_civil=data.get('id_estado_civil'),
            correo=data.get('correo'),
            domicilio=data.get('domicilio'),
            id_ciudad=data.get('id_ciudad'),
            id_ciudad_nacimiento=data.get('id_ciudad_nacimiento'),
            id_nivel_instruccion=data.get('id_nivel_instruccion'),
            id_profesion=data.get('id_profesion'),
            observaciones=data.get('observaciones'),
            nom_madre=data.get('nom_madre'),
            tel_madre=data.get('tel_madre'),
            nom_padre=data.get('nom_padre'),
            tel_padre=data.get('tel_padre'),
            educacion=data.get('educacion'),
            colegio=data.get('colegio'),
            tel_colegio=data.get('tel_colegio'),
            usuario_modificacion=session.get('id_usuario')
        )

        if resultado:
            return jsonify({
                'success': True,
                'data': {
                    'id_paciente': pac_id,
                    'mensaje': 'Paciente actualizado exitosamente'
                },
                'error': None
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': 'No se pudo actualizar el paciente.'
            }), 400

    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        app.logger.error(f"Error al actualizar paciente: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno.'
        }), 500


# ============================================
# DESACTIVAR PACIENTE
# ============================================
@pacienteapi.route('/pacientes/<int:pac_id>', methods=['DELETE'])
@role_required("ADMINISTRADOR", "SUPERADMIN")
def deletePaciente(pac_id):
    """Desactiva un paciente (soft-delete, no elimina datos)."""
    pacientedao = PacienteDao()

    try:
        if pacientedao.desactivarPaciente(pac_id, session.get('id_usuario')):
            return jsonify({
                'success': True,
                'mensaje': f'Paciente con ID {pac_id} desactivado correctamente.',
                'error': None
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': 'No se encontró el paciente con el ID proporcionado o no se pudo desactivar.'
            }), 404

    except Exception as e:
        app.logger.error(f"Error al desactivar paciente: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno. Consulte con el administrador.'
        }), 500
